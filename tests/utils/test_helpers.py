"""
共通テストユーティリティ
全テストで使用される共通機能
"""

import os
import sys
import json
import tempfile
from pathlib import Path
from typing import Dict, Any, Optional, List
from unittest.mock import MagicMock, patch
import logging

# プロジェクトルートをパスに追加
PROJECT_ROOT = Path(__file__).parent.parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

logger = logging.getLogger(__name__)


class TestDataProvider:
    """テストデータ提供クラス"""
    
    @staticmethod
    def get_sample_text() -> str:
        """サンプルテキストを取得"""
        return """
        一、次の文章を読んで、後の問いに答えなさい。
        
        私たちの生活は、多くの人々の支えによって成り立っている。
        朝起きてから夜寝るまで、私たちは数え切れないほどの
        恩恵を受けている。
        
        （山田太郎『日常の発見』による）
        
        問一　傍線部「恩恵」の意味として最も適切なものを選びなさい。
        ア　利益　　イ　感謝　　ウ　協力　　エ　贈り物
        
        問二　この文章の主題を三十字以内で説明しなさい。
        """
    
    @staticmethod
    def get_sample_pdf_path() -> Optional[Path]:
        """サンプルPDFパスを取得"""
        sample_dir = PROJECT_ROOT / "test_data"
        if sample_dir.exists():
            pdf_files = list(sample_dir.glob("*.pdf"))
            if pdf_files:
                return pdf_files[0]
        return None
    
    @staticmethod
    def get_sample_questions() -> List[Dict[str, Any]]:
        """サンプル設問データを取得"""
        return [
            {
                'number': 1,
                'type': '選択',
                'text': '傍線部「恩恵」の意味として最も適切なものを選びなさい。'
            },
            {
                'number': 2,
                'type': '記述',
                'text': 'この文章の主題を三十字以内で説明しなさい。'
            }
        ]
    
    @staticmethod
    def get_sample_school_data() -> Dict[str, Any]:
        """サンプル学校データを取得"""
        return {
            'name': '開成中学校',
            'year': 2025,
            'subject': '国語',
            'test_time': 50,
            'total_score': 100
        }


class MockFactory:
    """モックオブジェクト生成ファクトリー"""
    
    @staticmethod
    def create_mock_ocr_response() -> Dict[str, Any]:
        """OCRレスポンスのモックを作成"""
        return {
            'text': TestDataProvider.get_sample_text(),
            'confidence': 0.95,
            'language': 'ja',
            'pages': 1
        }
    
    @staticmethod
    def create_mock_config():
        """設定オブジェクトのモックを作成"""
        mock_config = MagicMock()
        mock_config.file.max_file_size_mb = 200
        mock_config.ocr.timeout_seconds = 300
        mock_config.processing.cache_enabled = False
        mock_config.paths.output_dir = Path("/tmp/test_output")
        return mock_config
    
    @staticmethod
    def create_mock_excel_file() -> Path:
        """一時的なExcelファイルのモックを作成"""
        import openpyxl
        
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "テストシート"
            ws['A1'] = "学校名"
            ws['B1'] = "年度"
            ws['A2'] = "開成中学校"
            ws['B2'] = 2025
            wb.save(tmp.name)
            return Path(tmp.name)


class FileTestHelper:
    """ファイル操作のテストヘルパー"""
    
    @staticmethod
    def create_temp_directory() -> Path:
        """一時ディレクトリを作成"""
        temp_dir = tempfile.mkdtemp(prefix="test_social_analyzer_")
        return Path(temp_dir)
    
    @staticmethod
    def create_temp_file(content: str, suffix: str = ".txt") -> Path:
        """一時ファイルを作成"""
        with tempfile.NamedTemporaryFile(mode='w', suffix=suffix, delete=False, encoding='utf-8') as tmp:
            tmp.write(content)
            return Path(tmp.name)
    
    @staticmethod
    def cleanup_temp_files(*paths: Path):
        """一時ファイルをクリーンアップ"""
        for path in paths:
            try:
                if path.is_file():
                    path.unlink()
                elif path.is_dir():
                    import shutil
                    shutil.rmtree(path)
            except Exception as e:
                logger.warning(f"Failed to cleanup {path}: {e}")


class AssertionHelper:
    """カスタムアサーションヘルパー"""
    
    @staticmethod
    def assert_valid_year(year: int):
        """年度の妥当性を検証"""
        assert 1990 <= year <= 2030, f"Invalid year: {year}"
    
    @staticmethod
    def assert_valid_school_name(name: str):
        """学校名の妥当性を検証"""
        valid_schools = [
            '開成中学校', '桜蔭中学校', '麻布中学校', 
            '武蔵中学校', '女子学院中学校', '雙葉中学校'
        ]
        assert any(school in name for school in valid_schools), f"Unknown school: {name}"
    
    @staticmethod
    def assert_valid_question_type(q_type: str):
        """設問タイプの妥当性を検証"""
        valid_types = ['選択', '記述', '抜き出し', '漢字・語句', '空欄補充', 'その他']
        assert q_type in valid_types, f"Invalid question type: {q_type}"
    
    @staticmethod
    def assert_section_structure(section: Dict[str, Any]):
        """セクション構造の妥当性を検証"""
        required_keys = ['number', 'questions', 'text']
        for key in required_keys:
            assert key in section, f"Missing required key in section: {key}"
        
        assert isinstance(section['number'], int), "Section number must be integer"
        assert isinstance(section['questions'], list), "Questions must be list"
        assert isinstance(section['text'], str), "Text must be string"


class PerformanceTimer:
    """パフォーマンス測定ヘルパー"""
    
    def __init__(self, name: str = "Operation"):
        self.name = name
        self.start_time = None
        self.end_time = None
    
    def __enter__(self):
        import time
        self.start_time = time.time()
        return self
    
    def __exit__(self, exc_type, exc_val, exc_tb):
        import time
        self.end_time = time.time()
        elapsed = self.end_time - self.start_time
        logger.info(f"{self.name} took {elapsed:.3f} seconds")
    
    @property
    def elapsed(self) -> float:
        """経過時間を取得"""
        if self.start_time and self.end_time:
            return self.end_time - self.start_time
        return 0.0


# テスト実行時の共通設定
def setup_test_environment():
    """テスト環境をセットアップ"""
    # ログ設定
    logging.basicConfig(
        level=logging.WARNING,
        format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
    )
    
    # 環境変数設定
    os.environ['ENTRANCE_EXAM_TEST_MODE'] = 'true'
    os.environ['ENTRANCE_EXAM_CACHE'] = 'false'
    
    # 一時ディレクトリ設定
    temp_dir = FileTestHelper.create_temp_directory()
    os.environ['ENTRANCE_EXAM_DATA_DIR'] = str(temp_dir)
    
    return temp_dir


def teardown_test_environment(temp_dir: Path):
    """テスト環境をクリーンアップ"""
    # 環境変数クリア
    for key in ['ENTRANCE_EXAM_TEST_MODE', 'ENTRANCE_EXAM_CACHE', 'ENTRANCE_EXAM_DATA_DIR']:
        if key in os.environ:
            del os.environ[key]
    
    # 一時ディレクトリ削除
    FileTestHelper.cleanup_temp_files(temp_dir)