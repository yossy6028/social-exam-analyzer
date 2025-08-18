"""
プラグインインターフェースパターン
拡張可能なプラグインシステムの基盤
"""

from abc import ABC, abstractmethod
from typing import Any, Dict, List, Optional, Type
from dataclasses import dataclass
import logging
from pathlib import Path
import importlib
import inspect

from core.config import MainConfig, get_config

logger = logging.getLogger(__name__)


@dataclass
class PluginMetadata:
    """プラグインメタデータ"""
    name: str
    version: str
    description: str
    author: str
    dependencies: List[str]
    config_schema: Optional[Dict[str, Any]] = None


class PluginInterface(ABC):
    """プラグインの基底インターフェース"""
    
    @abstractmethod
    def get_metadata(self) -> PluginMetadata:
        """プラグインメタデータを取得"""
        pass
    
    @abstractmethod
    def initialize(self, config: Dict[str, Any]) -> None:
        """
        プラグインを初期化
        
        Args:
            config: プラグイン設定
        """
        pass
    
    @abstractmethod
    def execute(self, data: Any, **kwargs) -> Any:
        """
        プラグインを実行
        
        Args:
            data: 入力データ
            **kwargs: 追加パラメータ
            
        Returns:
            処理結果
        """
        pass
    
    def validate_config(self, config: Dict[str, Any]) -> bool:
        """
        設定を検証
        
        Args:
            config: 検証する設定
            
        Returns:
            検証結果
        """
        metadata = self.get_metadata()
        if not metadata.config_schema:
            return True
        
        # 簡易的な検証（実際にはjsonschemaなどを使用）
        for key, schema in metadata.config_schema.items():
            if schema.get('required', False) and key not in config:
                logger.error(f"Required config key missing: {key}")
                return False
        
        return True
    
    def cleanup(self) -> None:
        """クリーンアップ処理"""
        pass


class InputPlugin(PluginInterface):
    """入力プラグインのインターフェース"""
    
    @abstractmethod
    def read(self, source: Any) -> Any:
        """
        データを読み込む
        
        Args:
            source: データソース
            
        Returns:
            読み込んだデータ
        """
        pass


class ProcessorPlugin(PluginInterface):
    """処理プラグインのインターフェース"""
    
    @abstractmethod
    def process(self, data: Any) -> Any:
        """
        データを処理
        
        Args:
            data: 入力データ
            
        Returns:
            処理済みデータ
        """
        pass


class OutputPlugin(PluginInterface):
    """出力プラグインのインターフェース"""
    
    @abstractmethod
    def write(self, data: Any, destination: Any) -> bool:
        """
        データを出力
        
        Args:
            data: 出力データ
            destination: 出力先
            
        Returns:
            成功かどうか
        """
        pass


class PluginManager:
    """プラグインマネージャー"""
    
    def __init__(self, config: Optional[MainConfig] = None):
        """
        初期化
        
        Args:
            config: アプリケーション設定
        """
        self.config = config or get_config()
        self.plugins: Dict[str, PluginInterface] = {}
        self.plugin_types: Dict[str, Type[PluginInterface]] = {
            'input': InputPlugin,
            'processor': ProcessorPlugin,
            'output': OutputPlugin
        }
    
    def register_plugin(
        self,
        plugin: PluginInterface,
        name: Optional[str] = None
    ) -> bool:
        """
        プラグインを登録
        
        Args:
            plugin: プラグインインスタンス
            name: プラグイン名（省略時はメタデータから取得）
            
        Returns:
            登録成功かどうか
        """
        try:
            metadata = plugin.get_metadata()
            plugin_name = name or metadata.name
            
            if plugin_name in self.plugins:
                logger.warning(f"Plugin {plugin_name} already registered, overwriting")
            
            self.plugins[plugin_name] = plugin
            logger.info(f"Registered plugin: {plugin_name} v{metadata.version}")
            return True
            
        except Exception as e:
            logger.error(f"Failed to register plugin: {e}")
            return False
    
    def unregister_plugin(self, name: str) -> bool:
        """
        プラグインを登録解除
        
        Args:
            name: プラグイン名
            
        Returns:
            成功かどうか
        """
        if name in self.plugins:
            plugin = self.plugins.pop(name)
            plugin.cleanup()
            logger.info(f"Unregistered plugin: {name}")
            return True
        return False
    
    def get_plugin(self, name: str) -> Optional[PluginInterface]:
        """
        プラグインを取得
        
        Args:
            name: プラグイン名
            
        Returns:
            プラグインインスタンス
        """
        return self.plugins.get(name)
    
    def list_plugins(self) -> List[PluginMetadata]:
        """
        登録済みプラグインのリストを取得
        
        Returns:
            プラグインメタデータのリスト
        """
        return [plugin.get_metadata() for plugin in self.plugins.values()]
    
    def load_plugin_from_file(self, file_path: Path) -> bool:
        """
        ファイルからプラグインを読み込む
        
        Args:
            file_path: プラグインファイルのパス
            
        Returns:
            成功かどうか
        """
        try:
            # モジュールを動的にインポート
            spec = importlib.util.spec_from_file_location(
                file_path.stem,
                file_path
            )
            module = importlib.util.module_from_spec(spec)
            spec.loader.exec_module(module)
            
            # プラグインクラスを探す
            for name, obj in inspect.getmembers(module):
                if (inspect.isclass(obj) and
                    issubclass(obj, PluginInterface) and
                    obj is not PluginInterface):
                    
                    # プラグインをインスタンス化して登録
                    plugin = obj()
                    self.register_plugin(plugin)
                    return True
            
            logger.warning(f"No plugin class found in {file_path}")
            return False
            
        except Exception as e:
            logger.error(f"Failed to load plugin from {file_path}: {e}")
            return False
    
    def load_plugins_from_directory(self, directory: Path) -> int:
        """
        ディレクトリからプラグインを一括読み込み
        
        Args:
            directory: プラグインディレクトリ
            
        Returns:
            読み込んだプラグイン数
        """
        loaded_count = 0
        
        if not directory.exists():
            logger.warning(f"Plugin directory does not exist: {directory}")
            return 0
        
        for file_path in directory.glob("*.py"):
            if file_path.name.startswith("_"):
                continue
            
            if self.load_plugin_from_file(file_path):
                loaded_count += 1
        
        logger.info(f"Loaded {loaded_count} plugins from {directory}")
        return loaded_count
    
    def execute_plugin(
        self,
        name: str,
        data: Any,
        **kwargs
    ) -> Optional[Any]:
        """
        プラグインを実行
        
        Args:
            name: プラグイン名
            data: 入力データ
            **kwargs: 追加パラメータ
            
        Returns:
            実行結果
        """
        plugin = self.get_plugin(name)
        if not plugin:
            logger.error(f"Plugin not found: {name}")
            return None
        
        try:
            return plugin.execute(data, **kwargs)
        except Exception as e:
            logger.error(f"Plugin execution failed: {e}")
            return None
    
    def create_pipeline(
        self,
        plugin_names: List[str]
    ) -> 'PluginPipeline':
        """
        プラグインパイプラインを作成
        
        Args:
            plugin_names: プラグイン名のリスト
            
        Returns:
            パイプラインインスタンス
        """
        plugins = []
        for name in plugin_names:
            plugin = self.get_plugin(name)
            if plugin:
                plugins.append(plugin)
            else:
                logger.warning(f"Plugin {name} not found, skipping")
        
        return PluginPipeline(plugins)


class PluginPipeline:
    """プラグインパイプライン"""
    
    def __init__(self, plugins: List[PluginInterface]):
        """
        初期化
        
        Args:
            plugins: プラグインのリスト
        """
        self.plugins = plugins
    
    def execute(self, data: Any, **kwargs) -> Any:
        """
        パイプラインを実行
        
        Args:
            data: 入力データ
            **kwargs: 追加パラメータ
            
        Returns:
            最終結果
        """
        result = data
        
        for plugin in self.plugins:
            try:
                metadata = plugin.get_metadata()
                logger.debug(f"Executing plugin: {metadata.name}")
                result = plugin.execute(result, **kwargs)
                
            except Exception as e:
                logger.error(f"Pipeline execution failed at {metadata.name}: {e}")
                raise
        
        return result


# 具体的なプラグイン実装例

class PDFInputPlugin(InputPlugin):
    """PDF入力プラグイン"""
    
    def get_metadata(self) -> PluginMetadata:
        return PluginMetadata(
            name="pdf_input",
            version="1.0.0",
            description="Read PDF files",
            author="System",
            dependencies=[]
        )
    
    def initialize(self, config: Dict[str, Any]) -> None:
        self.config = config
    
    def read(self, source: Path) -> str:
        """PDFを読み込む"""
        # 実際の実装では、PDFライブラリを使用
        logger.info(f"Reading PDF: {source}")
        return f"PDF content from {source}"
    
    def execute(self, data: Any, **kwargs) -> Any:
        return self.read(data)


class TextProcessorPlugin(ProcessorPlugin):
    """テキスト処理プラグイン"""
    
    def get_metadata(self) -> PluginMetadata:
        return PluginMetadata(
            name="text_processor",
            version="1.0.0",
            description="Process text data",
            author="System",
            dependencies=[]
        )
    
    def initialize(self, config: Dict[str, Any]) -> None:
        self.config = config
        from core.text_engine import TextEngine
        self.engine = TextEngine()
    
    def process(self, data: str) -> Dict[str, Any]:
        """テキストを処理"""
        processed = self.engine.process_text(data)
        return self.engine.export_to_dict(processed)
    
    def execute(self, data: Any, **kwargs) -> Any:
        return self.process(data)


class ExcelOutputPlugin(OutputPlugin):
    """Excel出力プラグイン"""
    
    def get_metadata(self) -> PluginMetadata:
        return PluginMetadata(
            name="excel_output",
            version="1.0.0",
            description="Write data to Excel",
            author="System",
            dependencies=["openpyxl"]
        )
    
    def initialize(self, config: Dict[str, Any]) -> None:
        self.config = config
    
    def write(self, data: Dict[str, Any], destination: Path) -> bool:
        """Excelに出力"""
        try:
            import openpyxl
            
            wb = openpyxl.Workbook()
            ws = wb.active
            
            # データを書き込み（簡易実装）
            ws['A1'] = "School"
            ws['B1'] = "Year"
            ws['C1'] = "Characters"
            
            ws['A2'] = data.get('school', '')
            ws['B2'] = data.get('year', '')
            ws['C2'] = data.get('total_characters', 0)
            
            wb.save(destination)
            logger.info(f"Saved to Excel: {destination}")
            return True
            
        except Exception as e:
            logger.error(f"Failed to write Excel: {e}")
            return False
    
    def execute(self, data: Any, **kwargs) -> Any:
        destination = kwargs.get('destination')
        if not destination:
            raise ValueError("destination is required")
        return self.write(data, destination)


# 使用例
def example_usage():
    """使用例"""
    # マネージャーを作成
    manager = PluginManager()
    
    # プラグインを登録
    manager.register_plugin(PDFInputPlugin())
    manager.register_plugin(TextProcessorPlugin())
    manager.register_plugin(ExcelOutputPlugin())
    
    # プラグインリストを表示
    for metadata in manager.list_plugins():
        print(f"- {metadata.name} v{metadata.version}: {metadata.description}")
    
    # パイプラインを作成
    pipeline = manager.create_pipeline([
        "pdf_input",
        "text_processor",
        "excel_output"
    ])
    
    # パイプラインを実行
    result = pipeline.execute(
        Path("sample.pdf"),
        destination=Path("output.xlsx")
    )
    
    print(f"Pipeline result: {result}")


if __name__ == "__main__":
    example_usage()