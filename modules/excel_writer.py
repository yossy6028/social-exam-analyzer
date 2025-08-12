"""
Excel出力モジュール
分析結果をExcelファイルに出力する
"""
import logging
from pathlib import Path
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


class ExcelWriter:
    """Excel出力クラス"""
    
    def __init__(self, output_dir: Path):
        """
        初期化
        
        Args:
            output_dir: 出力ディレクトリ
        """
        self.output_dir = output_dir
        
    def write_analysis_results(self, analysis_results: dict, 
                             school_name: str, year: str) -> Path:
        """
        分析結果をExcelに出力
        
        Args:
            analysis_results: 分析結果の辞書
            school_name: 学校名
            year: 年度
            
        Returns:
            出力ファイルのパス
        """
        # 出力ファイル名
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"{school_name}_{year}_分析結果_{timestamp}.xlsx"
        output_path = self.output_dir / filename
        
        # Workbookの作成
        wb = Workbook()
        
        # 既存のシートを削除
        wb.remove(wb.active)
        
        # 各シートの作成
        self._create_summary_sheet(wb, analysis_results, school_name, year)
        self._create_detail_sheet(wb, analysis_results)
        self._create_source_sheet(wb, analysis_results)
        
        # 保存
        wb.save(output_path)
        logger.info(f"分析結果を保存: {output_path}")
        
        return output_path
        
    def _create_summary_sheet(self, wb: Workbook, results: dict, 
                            school_name: str, year: str):
        """
        サマリーシートを作成
        """
        ws = wb.create_sheet('サマリー')
        
        # ヘッダー部分
        ws['A1'] = '国語入試問題分析結果'
        ws['A1'].font = Font(size=16, bold=True)
        
        ws['A3'] = '学校名:'
        ws['B3'] = school_name
        ws['A4'] = '年度:'
        ws['B4'] = year
        ws['A5'] = '分析日時:'
        ws['B5'] = datetime.now().strftime('%Y/%m/%d %H:%M')
        
        # 基本情報
        ws['A7'] = '■基本情報'
        ws['A7'].font = Font(bold=True)
        
        row = 8
        basic_info = [
            ('総文字数（概算）', results.get('total_characters', 0)),
            ('大問数', len(results.get('sections', []))),
            ('総設問数', len(results.get('questions', []))),
            ('文章テーマ', results.get('theme', '不明'))
        ]
        
        for label, value in basic_info:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            row += 1
            
        # 設問タイプ別集計
        row += 1
        ws[f'A{row}'] = '■設問タイプ別集計'
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        
        headers = ['設問タイプ', '問題数', '割合(%)']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='D3D3D3', 
                                  end_color='D3D3D3', 
                                  fill_type='solid')
        
        row += 1
        total_questions = len(results.get('questions', []))
        
        for q_type, count in results.get('question_types', {}).items():
            ws[f'A{row}'] = q_type
            ws[f'B{row}'] = count
            ws[f'C{row}'] = f"{(count / total_questions * 100):.1f}" if total_questions > 0 else "0"
            row += 1
            
        # 解答用紙情報（あれば）
        if 'answer_sheet_info' in results:
            row += 2
            ws[f'A{row}'] = '■解答用紙情報'
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
            
            answer_info = results['answer_sheet_info']
            if answer_info.get('total_points'):
                ws[f'A{row}'] = '満点'
                ws[f'B{row}'] = f"{answer_info['total_points']}点"
                row += 1
                
        # 列幅の調整
        for col in ['A', 'B', 'C']:
            ws.column_dimensions[col].width = 20
            
    def _create_detail_sheet(self, wb: Workbook, results: dict):
        """
        詳細分析シートを作成
        """
        ws = wb.create_sheet('詳細分析')
        
        # ヘッダー
        headers = ['大問', '設問番号', '設問マーカー', '設問タイプ', 
                  '選択肢数', '文字数制限', '配点']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='D3D3D3', 
                                  end_color='D3D3D3', 
                                  fill_type='solid')
            
        # データ部分
        row = 2
        for question in results.get('questions', []):
            ws.cell(row=row, column=1, value=question.get('section', ''))
            ws.cell(row=row, column=2, value=question.get('number', ''))
            ws.cell(row=row, column=3, value=question.get('marker', ''))
            ws.cell(row=row, column=4, value=question.get('type', ''))
            ws.cell(row=row, column=5, value=question.get('choice_count', ''))
            
            # 文字数制限
            char_limit = question.get('character_limit', '')
            if isinstance(char_limit, tuple):
                char_limit = f"{char_limit[0]}～{char_limit[1]}字"
            ws.cell(row=row, column=6, value=char_limit)
            
            # 配点（解答用紙情報から）
            point_key = f"問{question.get('number', '')}"
            points = ''
            if 'answer_sheet_info' in results:
                points = results['answer_sheet_info'].get('question_points', {}).get(point_key, '')
            ws.cell(row=row, column=7, value=points)
            
            row += 1
            
        # 列幅の調整
        for col in range(1, 8):
            ws.column_dimensions[get_column_letter(col)].width = 15
            
    def _create_source_sheet(self, wb: Workbook, results: dict):
        """
        出典一覧シートを作成
        """
        ws = wb.create_sheet('出典一覧')
        
        # ヘッダー
        headers = ['大問', '著者', '作品名', '出版社', '出版年', '備考']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='D3D3D3', 
                                  end_color='D3D3D3', 
                                  fill_type='solid')
            
        # データ部分
        row = 2
        for i, section in enumerate(results.get('sections', []), 1):
            source_info = section.get('source_info', {})
            
            ws.cell(row=row, column=1, value=f"大問{i}")
            ws.cell(row=row, column=2, value=source_info.get('author', ''))
            ws.cell(row=row, column=3, value=source_info.get('title', ''))
            ws.cell(row=row, column=4, value=source_info.get('publisher', ''))
            ws.cell(row=row, column=5, value=source_info.get('year', ''))
            ws.cell(row=row, column=6, value=source_info.get('raw_source', ''))
            
            row += 1
            
        # 列幅の調整
        column_widths = [10, 20, 30, 20, 10, 40]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width