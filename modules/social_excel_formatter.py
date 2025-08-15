"""
社会科目分析結果のExcel出力モジュール
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from typing import Dict, List, Any
import logging
from datetime import datetime
from .social_analyzer import SocialQuestion

logger = logging.getLogger(__name__)


class SocialExcelFormatter:
    """社会科目分析結果をExcel形式で出力"""
    
    def __init__(self):
        self.wb = None
        self.setup_styles()
    
    def setup_styles(self):
        """Excel用のスタイル設定"""
        self.header_font = Font(bold=True, size=12)
        self.title_font = Font(bold=True, size=14)
        self.header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        self.header_white_font = Font(bold=True, color="FFFFFF", size=11)
        
        # 分野別の色設定
        self.field_colors = {
            "地理": "E8F4EA",  # 薄緑
            "歴史": "FFF4E6",  # 薄オレンジ
            "公民": "E6F3FF",  # 薄青
            "時事問題": "FFE6E6",  # 薄赤
            "総合": "F0F0F0",  # 薄灰
        }
        
        # 枠線設定
        self.thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def create_excel_report(self, analysis_results: Dict[str, Any], school_name: str, year: str) -> Workbook:
        """分析結果からExcelレポートを作成"""
        self.wb = Workbook()
        
        # 既存のシートを削除
        if 'Sheet' in self.wb.sheetnames:
            self.wb.remove(self.wb['Sheet'])
        
        # 各シートを作成
        self._create_summary_sheet(analysis_results['statistics'], school_name, year)
        self._create_questions_sheet(analysis_results['questions'])
        self._create_field_analysis_sheet(analysis_results['statistics'])
        self._create_resource_analysis_sheet(analysis_results['statistics'])
        self._create_format_analysis_sheet(analysis_results['statistics'])
        self._create_current_affairs_sheet(analysis_results['questions'])
        
        return self.wb
    
    def _create_summary_sheet(self, stats: Dict, school_name: str, year: str):
        """概要シートの作成"""
        ws = self.wb.create_sheet("概要")
        
        # タイトル
        ws['A1'] = f"{school_name} {year}年度 社会科入試問題分析"
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:F1')
        
        # 基本情報
        ws['A3'] = "分析日時"
        ws['B3'] = datetime.now().strftime("%Y年%m月%d日 %H:%M")
        
        row = 5
        ws[f'A{row}'] = "【分野別出題状況】"
        ws[f'A{row}'].font = self.title_font
        row += 1
        
        # 分野別統計
        if 'field_distribution' in stats:
            for field, data in stats['field_distribution'].items():
                ws[f'A{row}'] = field
                ws[f'B{row}'] = f"{data['count']}問"
                ws[f'C{row}'] = f"{data['percentage']}%"
                
                # 背景色を設定
                if field in self.field_colors:
                    fill = PatternFill(start_color=self.field_colors[field], 
                                     end_color=self.field_colors[field], 
                                     fill_type="solid")
                    for col in ['A', 'B', 'C']:
                        ws[f'{col}{row}'].fill = fill
                row += 1
        
        row += 1
        ws[f'A{row}'] = "【資料活用状況】"
        ws[f'A{row}'].font = self.title_font
        row += 1
        
        if 'has_resources' in stats:
            ws[f'A{row}'] = "資料あり問題"
            ws[f'B{row}'] = f"{stats['has_resources']['count']}問"
            ws[f'C{row}'] = f"{stats['has_resources']['percentage']}%"
            row += 1
        
        row += 1
        ws[f'A{row}'] = "【時事問題】"
        ws[f'A{row}'].font = self.title_font
        row += 1
        
        if 'current_affairs' in stats:
            ws[f'A{row}'] = "時事問題"
            ws[f'B{row}'] = f"{stats['current_affairs']['count']}問"
            ws[f'C{row}'] = f"{stats['current_affairs']['percentage']}%"
        
        # 列幅調整
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
    
    def _create_questions_sheet(self, questions: List[SocialQuestion]):
        """問題詳細シートの作成"""
        ws = self.wb.create_sheet("問題詳細")
        
        # ヘッダー行
        headers = ["問題番号", "分野", "資料種別", "出題形式", "時事", "時代/地域", "主題", "キーワード", "問題文（冒頭）"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_white_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.thin_border
        
        # データ行
        for row, q in enumerate(questions, 2):
            ws.cell(row=row, column=1, value=q.number)
            ws.cell(row=row, column=2, value=q.field.value)
            ws.cell(row=row, column=3, value=", ".join([r.value for r in q.resource_types]))
            ws.cell(row=row, column=4, value=q.question_format.value)
            ws.cell(row=row, column=5, value="○" if q.is_current_affairs else "")
            ws.cell(row=row, column=6, value=q.time_period or q.region or "")
            # 公民の凡庸テーマは具体化（条文優先）
            topic_display = q.topic or ""
            try:
                if (not topic_display or topic_display in ['日本国憲法の原則','政治の仕組み','経済の仕組み']) and q.field.value == '公民':
                    import re
                    arts = re.findall(r'第(\d+)条', q.text)
                    if arts:
                        topic_display = f"憲法第{arts[0]}条"
            except Exception:
                pass
            ws.cell(row=row, column=7, value=topic_display)
            ws.cell(row=row, column=8, value=", ".join(q.keywords[:5]))
            ws.cell(row=row, column=9, value=q.text[:100] + "..." if len(q.text) > 100 else q.text)
            
            # 分野別背景色
            if q.field.value in self.field_colors:
                fill = PatternFill(start_color=self.field_colors[q.field.value],
                                 end_color=self.field_colors[q.field.value],
                                 fill_type="solid")
                ws.cell(row=row, column=2).fill = fill
            
            # 枠線
            for col in range(1, 10):
                ws.cell(row=row, column=col).border = self.thin_border
        
        # 列幅調整
        column_widths = [12, 10, 20, 12, 8, 15, 20, 30, 50]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[chr(64 + col)].width = width
    
    def _create_field_analysis_sheet(self, stats: Dict):
        """分野別分析シートの作成"""
        ws = self.wb.create_sheet("分野別分析")
        
        ws['A1'] = "分野別出題分析"
        ws['A1'].font = self.title_font
        ws.merge_cells('A1:D1')
        
        # ヘッダー
        headers = ["分野", "出題数", "割合(%)", "グラフ"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = self.header_white_font
            cell.fill = self.header_fill
            cell.border = self.thin_border
        
        # データ
        row = 4
        if 'field_distribution' in stats:
            for field, data in stats['field_distribution'].items():
                ws.cell(row=row, column=1, value=field)
                ws.cell(row=row, column=2, value=data['count'])
                ws.cell(row=row, column=3, value=data['percentage'])
                
                # 簡易グラフ（バー）
                bar_length = int(data['percentage'] / 5)
                ws.cell(row=row, column=4, value="■" * bar_length)
                
                # 背景色
                if field in self.field_colors:
                    fill = PatternFill(start_color=self.field_colors[field],
                                     end_color=self.field_colors[field],
                                     fill_type="solid")
                    for col in range(1, 5):
                        ws.cell(row=row, column=col).fill = fill
                        ws.cell(row=row, column=col).border = self.thin_border
                
                row += 1
        
        # 列幅調整
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 10
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 25
    
    def _create_resource_analysis_sheet(self, stats: Dict):
        """資料活用分析シートの作成"""
        ws = self.wb.create_sheet("資料活用分析")
        
        ws['A1'] = "資料活用状況分析"
        ws['A1'].font = self.title_font
        ws.merge_cells('A1:D1')
        
        # ヘッダー
        headers = ["資料種別", "使用回数", "割合(%)", "グラフ"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = self.header_white_font
            cell.fill = self.header_fill
            cell.border = self.thin_border
        
        # データ
        row = 4
        if 'resource_usage' in stats:
            # 使用回数でソート
            sorted_resources = sorted(stats['resource_usage'].items(), 
                                    key=lambda x: x[1]['count'], 
                                    reverse=True)
            
            for resource, data in sorted_resources:
                ws.cell(row=row, column=1, value=resource)
                ws.cell(row=row, column=2, value=data['count'])
                ws.cell(row=row, column=3, value=data['percentage'])
                
                # 簡易グラフ
                bar_length = int(data['percentage'] / 5)
                ws.cell(row=row, column=4, value="■" * bar_length)
                
                # 枠線
                for col in range(1, 5):
                    ws.cell(row=row, column=col).border = self.thin_border
                
                row += 1
        
        # 列幅調整
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 10
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 25
    
    def _create_format_analysis_sheet(self, stats: Dict):
        """出題形式分析シートの作成"""
        ws = self.wb.create_sheet("出題形式分析")
        
        ws['A1'] = "出題形式別分析"
        ws['A1'].font = self.title_font
        ws.merge_cells('A1:D1')
        
        # ヘッダー
        headers = ["出題形式", "出題数", "割合(%)", "グラフ"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = self.header_white_font
            cell.fill = self.header_fill
            cell.border = self.thin_border
        
        # データ
        row = 4
        if 'format_distribution' in stats:
            # 出題数でソート
            sorted_formats = sorted(stats['format_distribution'].items(),
                                   key=lambda x: x[1]['count'],
                                   reverse=True)
            
            for format_type, data in sorted_formats:
                ws.cell(row=row, column=1, value=format_type)
                ws.cell(row=row, column=2, value=data['count'])
                ws.cell(row=row, column=3, value=data['percentage'])
                
                # 簡易グラフ
                bar_length = int(data['percentage'] / 5)
                ws.cell(row=row, column=4, value="■" * bar_length)
                
                # 枠線
                for col in range(1, 5):
                    ws.cell(row=row, column=col).border = self.thin_border
                
                row += 1
        
        # 列幅調整
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 10
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 25
    
    def _create_current_affairs_sheet(self, questions: List[SocialQuestion]):
        """時事問題シートの作成"""
        ws = self.wb.create_sheet("時事問題")
        
        ws['A1'] = "時事問題詳細"
        ws['A1'].font = self.title_font
        ws.merge_cells('A1:E1')
        
        # ヘッダー
        headers = ["問題番号", "分野", "主題", "キーワード", "問題文（抜粋）"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = self.header_white_font
            cell.fill = self.header_fill
            cell.border = self.thin_border
        
        # 時事問題のみ抽出
        current_affairs_questions = [q for q in questions if q.is_current_affairs]
        
        if not current_affairs_questions:
            ws.cell(row=4, column=1, value="時事問題は検出されませんでした")
        else:
            row = 4
            for q in current_affairs_questions:
                ws.cell(row=row, column=1, value=q.number)
                ws.cell(row=row, column=2, value=q.field.value)
                ws.cell(row=row, column=3, value=q.topic or "")
                ws.cell(row=row, column=4, value=", ".join(q.keywords[:5]))
                ws.cell(row=row, column=5, value=q.text[:150] + "..." if len(q.text) > 150 else q.text)
                
                # 枠線
                for col in range(1, 6):
                    ws.cell(row=row, column=col).border = self.thin_border
                
                row += 1
        
        # 列幅調整
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 10
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 30
        ws.column_dimensions['E'].width = 50
    
    def save_excel(self, wb: Workbook, filepath: str):
        """Excelファイルを保存"""
        try:
            wb.save(filepath)
            logger.info(f"Excelファイルを保存しました: {filepath}")
        except Exception as e:
            logger.error(f"Excelファイルの保存に失敗しました: {e}")
            raise